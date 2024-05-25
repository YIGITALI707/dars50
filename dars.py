import openpyxl
wb = openpyxl.load_workbook('users.xlsx')
fayl= wb.active
fayl2= []
for i in fayl.iter_rows(values_only=True):
    fayl2.append(i)
fayl3= openpyxl.Workbook()
fayl4= fayl3.active
for x in fayl2:
    fayl4.append(x)
fayl5= 'users_table.xlsx'
fayl3.save(fayl5)
print(fayl5)