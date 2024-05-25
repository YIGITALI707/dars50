import openpyxl
a=openpyxl.load_workbook('users.xlsx')
sheet =a.active
fayl=[]
for i in sheet.iter_rows(values_only=True):
     fayl.append(i)
fayl1= openpyxl.Workbook()
fayl2= fayl1.active
for x in fayl:
    fayl2.append(x)
fayl4='users_table.xlsx'
fayl1.save(fayl4)
print(f'{fayl4} saqlash.')
